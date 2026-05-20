from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import anthropic
import os
import json
import requests as req_lib

app = Flask(__name__, static_folder='.')
CORS(app)

ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY")
NOTION_TOKEN  = os.environ.get("NOTION_TOKEN")
DATA_FILE     = "queue.json"

client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

def load_queue():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_queue(queue):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(queue, f, ensure_ascii=False, indent=2)

def notion_headers():
    return {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json"
    }

def extract_page_id(url):
    pid = url.split("?")[0].split("/")[-1].replace("-", "")
    return pid[-32:] if len(pid) >= 32 else pid

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/health")
def health():
    return jsonify({"status": "ok"})

# ── 노션 페이지 리딩 ──────────────────────────────────
@app.route("/notion-read", methods=["POST"])
def notion_read():
    try:
        data = request.json
        notion_url = data.get("notion_url", "")
        if not notion_url:
            return jsonify({"ok": False, "error": "URL 없음"})
        if not NOTION_TOKEN:
            return jsonify({"ok": False, "error": "NOTION_TOKEN 미설정"})

        page_id = extract_page_id(notion_url)

        # 페이지 속성 조회
        page_resp = req_lib.get(
            f"https://api.notion.com/v1/pages/{page_id}",
            headers=notion_headers()
        )
        if not page_resp.ok:
            return jsonify({"ok": False, "error": page_resp.text})

        page_data = page_resp.json()
        props = page_data.get("properties", {})

        def get_text(prop):
            if not prop: return ""
            t = prop.get("type")
            if t == "rich_text":
                return "".join(r.get("plain_text","") for r in prop.get("rich_text",[]))
            if t == "title":
                return "".join(r.get("plain_text","") for r in prop.get("title",[]))
            if t == "number":
                v = prop.get("number")
                return str(int(v)) if v is not None else ""
            if t == "select":
                s = prop.get("select")
                return s.get("name","") if s else ""
            if t == "date":
                d = prop.get("date")
                return d.get("start","") if d else ""
            if t == "files":
                names = []
                for f in prop.get("files",[]):
                    n = f.get("name","")
                    if n: names.append(n)
                return ", ".join(names)
            return ""

        buyer_name = get_text(props.get("업체명"))
        total_vat  = get_text(props.get("최종매출액(vat포함)"))
        file_names = get_text(props.get("견적서/사업자등록증"))

        # 블록(본문) 조회 → 계산서 발행 요청 섹션 파싱
        blocks_resp = req_lib.get(
            f"https://api.notion.com/v1/blocks/{page_id}/children?page_size=100",
            headers=notion_headers()
        )

        issue_date = ""
        recv_email = ""
        item_text  = ""

        if blocks_resp.ok:
            blocks = blocks_resp.json().get("results", [])
            in_section = False
            for b in blocks:
                btype = b.get("type","")
                rich  = b.get(btype,{}).get("rich_text",[]) if btype != "heading_2" else b.get("heading_2",{}).get("rich_text",[])
                text  = "".join(r.get("plain_text","") for r in rich)

                if "계산서 발행 요청" in text:
                    in_section = True
                    continue

                if in_section:
                    # 다음 heading이 나오면 섹션 종료
                    if btype in ("heading_1","heading_2","heading_3") and "계산서" not in text:
                        in_section = False
                        continue

                    tl = text.lower()
                    if "발급일" in text and not issue_date:
                        # "발급일: 5/20" or "발급일 : 26.05.20"
                        parts = text.replace("발급일","").replace("(업체 요청 시)","").replace(":","").replace("**","").strip()
                        issue_date = parts.strip()

                    if "이메일" in text and not recv_email:
                        import re
                        emails = re.findall(r"[\w\.-]+@[\w\.-]+\.\w+", text)
                        if emails:
                            recv_email = emails[0]

                    if "품목" in text and not item_text:
                        parts = text.replace("품목","").replace("(업체 요청 시)","").replace(":","").replace("**","").strip()
                        if parts and "미기재" not in parts and len(parts) > 1:
                            item_text = parts.strip()

        return jsonify({
            "ok": True,
            "buyer_name": buyer_name,
            "total_vat": total_vat,
            "file_names": file_names,
            "issue_date_raw": issue_date,
            "recv_email": recv_email,
            "item_text": item_text,
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ── AI Extract ────────────────────────────────────────
@app.route("/extract", methods=["POST"])
def extract():
    try:
        data      = request.json
        biz_b64   = data.get("biz_b64")
        biz_mt    = data.get("biz_mt", "image/jpeg")
        quote_b64 = data.get("quote_b64")
        quote_mt  = data.get("quote_mt", "image/jpeg")
        hint      = data.get("hint", "")  # 노션에서 읽은 힌트 정보

        def make_block(b64, mt):
            if mt == "application/pdf":
                return {"type": "document", "source": {"type": "base64", "media_type": mt, "data": b64}}
            return {"type": "image", "source": {"type": "base64", "media_type": mt, "data": b64}}

        hint_text = f"\n\n참고 정보 (노션에서 읽음): {hint}" if hint else ""

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{"role": "user", "content": [
                make_block(biz_b64, biz_mt),
                make_block(quote_b64, quote_mt),
                {"type": "text", "text": f"""첫 번째 문서는 공급받는자의 사업자등록증 또는 고유번호증, 두 번째는 견적서입니다.{hint_text}

JSON만 출력하세요:
{{"buyer_reg":"","buyer_name":"","buyer_ceo":"","buyer_addr":"","buyer_biz":"","buyer_item":"","buyer_email":"","issue_date":"YYYYMMDD","total_supply":"숫자만","total_tax":"숫자만","total_vat_included":"VAT포함총액숫자만","item_summary":"품목통칭(예:종이구조물 외)","items":[{{"day":"DD","name":"","spec":"","qty":"","price":"","supply":"","tax":""}}]}}
모르는 값은 빈 문자열. items 최대 4개. item_summary는 품목이 여러 개면 대표품목 뒤에 '외'를 붙여 한 줄로."""}
            ]}]
        )

        raw    = next((b.text for b in response.content if b.type == "text"), "")
        parsed = json.loads(raw.replace("```json", "").replace("```", "").strip())
        return jsonify({"ok": True, "data": parsed})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ── Queue CRUD ────────────────────────────────────────
@app.route("/queue", methods=["POST"])
def add_entry():
    try:
        entry = request.json.get("entry")
        queue = load_queue()
        queue.append(entry)
        save_queue(queue)
        return jsonify({"ok": True, "count": len(queue)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/queue", methods=["GET"])
def get_queue():
    try:
        return jsonify({"ok": True, "data": load_queue()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/queue/<int:entry_id>", methods=["DELETE"])
def delete_entry(entry_id):
    try:
        queue = [e for e in load_queue() if e.get("id") != entry_id]
        save_queue(queue)
        return jsonify({"ok": True, "count": len(queue)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/queue/<int:entry_id>", methods=["PATCH"])
def update_entry(entry_id):
    try:
        updates = request.json
        queue   = load_queue()
        for e in queue:
            if e.get("id") == entry_id:
                e.update(updates)
                break
        save_queue(queue)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ── 발급 완료 → 노션 기입 ─────────────────────────────
@app.route("/notion-complete", methods=["POST"])
def notion_complete():
    try:
        data       = request.json
        notion_url = data.get("notion_url", "")
        issue_date = data.get("issue_date", "")   # YYYY-MM-DD
        buyer_name = data.get("buyer_name", "")
        total_vat  = data.get("total_vat", "")

        if not NOTION_TOKEN:
            return jsonify({"ok": False, "error": "NOTION_TOKEN 미설정"})
        if not notion_url:
            return jsonify({"ok": False, "error": "노션 URL 없음"})

        page_id = extract_page_id(notion_url)

        properties = {}

        if buyer_name:
            properties["업체명"] = {
                "rich_text": [{"type": "text", "text": {"content": buyer_name}}]
            }

        if total_vat:
            try:
                properties["최종매출액(vat포함)"] = {"number": int(str(total_vat).replace(",","").strip())}
            except:
                properties["최종매출액(vat포함)"] = {
                    "rich_text": [{"type": "text", "text": {"content": str(total_vat)}}]
                }

        if issue_date:
            # YYYYMMDD → YYYY-MM-DD
            d = str(issue_date).replace("-","")
            if len(d) == 8:
                issue_date = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
            properties["[★]계산서발행일/결제일"] = {"date": {"start": issue_date}}

        resp = req_lib.patch(
            f"https://api.notion.com/v1/pages/{page_id}",
            headers=notion_headers(),
            json={"properties": properties}
        )

        if resp.ok:
            return jsonify({"ok": True})
        else:
            return jsonify({"ok": False, "error": resp.text}), 500

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)


# ── 국세청 양식 코드로 직접 생성 ─────────────────────
@app.route("/export-excel", methods=["POST"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        from flask import send_file
        from datetime import date

        data     = request.json
        entries  = data.get("entries", [])
        supplier = data.get("supplier", {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "엑셀업로드양식"

        def to_num(v):
            try: return int(str(v).replace(",","").strip()) if v else None
            except: return str(v) if v else None

        # 헤더 (6행)
        headers = [
            '전자(세금)계산서 종류\n(01:일반, 02:영세율)', '작성일자',
            '공급자 등록번호\n("-" 없이 입력)', '공급자\n 종사업장번호',
            '공급자 상호', '공급자 성명', '공급자 사업장주소', '공급자 업태', '공급자 종목', '공급자 이메일',
            '공급받는자 등록번호\n("-" 없이 입력)', '공급받는자 \n종사업장번호',
            '공급받는자 상호 ', '공급받는자 성명', '공급받는자 사업장주소',
            '공급받는자 업태', '공급받는자 종목', '공급받는자 이메일1', '공급받는자 이메일2',
            '공급가액\n합계', '세액\n합계', '비고',
            '일자1\n(2자리, 작성년월 제외)', '품목1', '규격1', '수량1', '단가1', '공급가액1', '세액1', '품목비고1',
            '일자2\n(2자리, 작성년월 제외)', '품목2', '규격2', '수량2', '단가2', '공급가액2', '세액2', '품목비고2',
            '일자3\n(2자리, 작성년월 제외)', '품목3', '규격3', '수량3', '단가3', '공급가액3', '세액3', '품목비고3',
            '일자4\n(2자리, 작성년월 제외)', '품목4', '규격4', '수량4', '단가4', '공급가액4', '세액4', '품목비고4',
            '현금', '수표', '어음', '외상미수금', '영수(01),\n청구(02)'
        ]

        # 헤더 스타일
        header_fill = PatternFill("solid", fgColor="FFA500")
        header_font = Font(bold=True, size=9)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        ws.row_dimensions[6].height = 40

        # 데이터 (7행~)
        for i, e in enumerate(entries):
            row = 7 + i
            its = (e.get("items") or []) + [{}] * 4
            its = its[:4]

            vals = [
                "01", e.get("issue_date",""),
                supplier.get("reg",""), None,
                supplier.get("name",""), supplier.get("ceo",""),
                supplier.get("addr",""), supplier.get("biz",""),
                supplier.get("item",""), supplier.get("email",""),
                e.get("buyer_reg",""), None,
                e.get("buyer_name",""), e.get("buyer_ceo",""),
                e.get("buyer_addr",""), e.get("buyer_biz",""),
                e.get("buyer_item",""), e.get("buyer_email",""), None,
                to_num(e.get("total_supply")), to_num(e.get("total_tax")),
                e.get("note",""),
                its[0].get("day",""), its[0].get("name",""), its[0].get("spec",""), to_num(its[0].get("qty")), to_num(its[0].get("price")), to_num(its[0].get("supply")), to_num(its[0].get("tax")), None,
                its[1].get("day",""), its[1].get("name",""), its[1].get("spec",""), to_num(its[1].get("qty")), to_num(its[1].get("price")), to_num(its[1].get("supply")), to_num(its[1].get("tax")), None,
                its[2].get("day",""), its[2].get("name",""), its[2].get("spec",""), to_num(its[2].get("qty")), to_num(its[2].get("price")), to_num(its[2].get("supply")), to_num(its[2].get("tax")), None,
                its[3].get("day",""), its[3].get("name",""), its[3].get("spec",""), to_num(its[3].get("qty")), to_num(its[3].get("price")), to_num(its[3].get("supply")), to_num(its[3].get("tax")), None,
                None, None, None, None,
                e.get("receipt","02"),
            ]

            for col, val in enumerate(vals, 1):
                if val is not None and val != "":
                    ws.cell(row=row, column=col, value=val)

        # 컬럼 너비
        col_widths = [8,12,16,8,16,10,30,12,16,20,16,8,16,10,30,12,16,20,20,12,10,12] + [6,16,8,6,10,10,8,8]*4 + [8,8,8,10,8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        fname = f"세금계산서_{date.today().strftime('%Y%m%d')}_{len(entries)}건.xlsx"
        return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=fname)

    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500
